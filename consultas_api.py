from flask import Blueprint, request, jsonify, Response, send_from_directory
from db import get_connection
from io import StringIO, BytesIO
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from datetime import datetime, timezone, timedelta
from flask import send_file
import os
from logs_db import registrar_log 

def ms_a_fecha_local(ms):
    if not ms:
        return ""
    dt_utc = datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc)
    dt_local = dt_utc.astimezone(tz=timezone(timedelta(hours=-5)))  # 🇨🇴 Colombia
    return dt_local.strftime("%d/%m/%Y %H:%M:%S")

consultas_bp = Blueprint("consultas", __name__)

@consultas_bp.route("/api/tunnels", methods=["GET"])
def listar_tuneles():
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name, created_at 
            FROM tunnels 
            ORDER BY id DESC
        """)
        tuneles = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(tuneles)
    except Exception as e:
        print("❌ Error listando túneles:", e)
        return jsonify({"error": "Error interno"}), 500

@consultas_bp.route("/api/messages", methods=["GET"])
def listar_mensajes():
    tunnel_id = request.args.get("tunnel_id")
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")

    if not tunnel_id:
        return jsonify({"error": "Falta el parámetro tunnel_id"}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        query = """
            SELECT id, client_uuid, alias, contenido, tipo, enviado_en
            FROM tunnel_messages
            WHERE tunnel_id = %s
        """
        params = [tunnel_id]

        if desde:
            query += " AND enviado_en >= %s"
            params.append(desde)
        if hasta:
            query += " AND enviado_en <= %s"
            params.append(hasta)

        query += " ORDER BY enviado_en ASC"

        cursor.execute(query, tuple(params))
        mensajes = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(mensajes)
    except Exception as e:
        print("❌ Error listando mensajes:", e)
        return jsonify({"error": "Error interno"}), 500

@consultas_bp.route("/api/files", methods=["GET"])
def listar_archivos():
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT f.id, f.filename, f.uploaded_at, f.tunnel_id, f.client_uuid, f.sender_alias, t.name as tunnel_name
            FROM tunnel_files f
            LEFT JOIN tunnels t ON t.id = f.tunnel_id
            ORDER BY f.id DESC
        """)
        archivos = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(archivos)
    except Exception as e:
        print("❌ Error listando archivos:", e)
        return jsonify({"error": "Error interno"}), 500


@consultas_bp.route("/api/users", methods=["GET"])
def listar_usuarios():
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT DISTINCT client_uuid AS uuid, alias
            FROM client_aliases
            ORDER BY alias
        """)
        usuarios = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(usuarios)
    except Exception as e:
        print("❌ Error listando usuarios:", e)
        return jsonify({"error": "Error interno"}), 500

@consultas_bp.route("/api/files_by_day", methods=["GET"])
def archivos_por_dia():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DATE(uploaded_at) as fecha, COUNT(*) as total
            FROM tunnel_files
            GROUP BY DATE(uploaded_at)
            ORDER BY fecha
        """)
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify([
            {"fecha": str(row[0]), "total": row[1]}
            for row in resultados
        ])
    except Exception as e:
        print("❌ Error agrupando archivos:", e)
        return jsonify({"error": "Error interno"}), 500

@consultas_bp.route("/api/clientes", methods=["GET"])
def listar_clientes():
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT uuid, hostname, sistema_operativo, creado_en
            FROM clients
            ORDER BY creado_en DESC
        """)
        clientes = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(clientes)
    except Exception as e:
        print("❌ Error listando clientes:", e)
        return jsonify({"error": "Error interno"}), 500

@consultas_bp.route("/api/tunnels/<int:tunnel_id>/participantes", methods=["GET"])
def listar_participantes(tunnel_id):
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT ca.client_uuid, cl.hostname, ca.alias, MAX(ca.conectado_en) as ultimo_acceso
            FROM client_aliases ca
            LEFT JOIN clients cl ON cl.uuid = ca.client_uuid
            WHERE ca.tunnel_id = %s
            GROUP BY ca.client_uuid, cl.hostname, ca.alias
        """, (tunnel_id,))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        agrupado = {}
        for row in rows:
            clave = (row["client_uuid"], row["hostname"])
            if clave not in agrupado:
                agrupado[clave] = {
                    "client_uuid": row["client_uuid"],
                    "hostname": row["hostname"],
                    "aliases": set(),
                    "ultimo_acceso": row["ultimo_acceso"]
                }
            agrupado[clave]["aliases"].add(row["alias"])
            if row["ultimo_acceso"] > agrupado[clave]["ultimo_acceso"]:
                agrupado[clave]["ultimo_acceso"] = row["ultimo_acceso"]

        return jsonify([
            {
                "client_uuid": val["client_uuid"],
                "hostname": val["hostname"],
                "aliases": list(val["aliases"]),
                "ultimo_acceso": val["ultimo_acceso"]
            }
            for val in agrupado.values()
        ])
    except Exception as e:
        print("❌ Error agrupando participantes:", e)
        return jsonify({"error": "Error interno"}), 500

@consultas_bp.route("/api/tunnels/<int:tunnel_id>/export", methods=["GET"])
def exportar_chat(tunnel_id):
    formato = request.args.get("formato", "csv")
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    username = request.args.get("username")  # ← Captura del usuario para el log

    try:
        zona_colombia = timezone(timedelta(hours=-5))
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT name FROM tunnels WHERE id = %s", (tunnel_id,))
        tunel_info = cursor.fetchone()
        tunel_nombre = tunel_info["name"] if tunel_info else f"tunel_{tunnel_id}"

        query = """
            SELECT alias, contenido, enviado_en 
            FROM tunnel_messages
            WHERE tunnel_id = %s
        """
        params = [tunnel_id]

        if desde:
            query += " AND enviado_en >= %s"
            params.append(int(desde))
        if hasta:
            query += " AND enviado_en <= %s"
            params.append(int(hasta))

        query += " ORDER BY id ASC"
        cursor.execute(query, tuple(params))
        mensajes = cursor.fetchall()
        cursor.close()
        conn.close()

        if not mensajes:
            print(f"⚠️ No hay mensajes para tunel_id={tunnel_id} en el rango {desde} - {hasta}")

        # ✅ REGISTRO DEL LOG
        if username:
            from logs_db import registrar_log
            ip = request.remote_addr
            registrar_log(username, f"Descargó chat del túnel '{tunel_nombre}' (ID {tunnel_id})", "Exportación", ip)

        if formato == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(["Alias", "Contenido", "Fecha"])
            for msg in mensajes:
                dt_local = datetime.fromtimestamp(msg["enviado_en"] / 1000, tz=timezone.utc).astimezone(zona_colombia)
                fecha_legible = dt_local.strftime("%Y-%m-%d %H:%M:%S")
                writer.writerow([msg["alias"], msg["contenido"], fecha_legible])
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype="text/csv",
                headers={"Content-Disposition": f"attachment; filename=chat_{tunel_nombre}.csv"}
            )

        elif formato == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{tunel_nombre}"
            ws.append(["Alias", "Contenido", "Fecha"])
            for msg in mensajes:
                dt_local = datetime.fromtimestamp(msg["enviado_en"] / 1000, tz=timezone.utc).astimezone(zona_colombia)
                fecha_legible = dt_local.strftime("%Y-%m-%d %H:%M:%S")
                ws.append([msg["alias"], msg["contenido"], fecha_legible])
            for i in range(1, 4):
                ws.column_dimensions[get_column_letter(i)].width = 30
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=chat_{tunel_nombre}.xlsx"}
            )

        return jsonify({"error": "Formato no soportado"}), 400

    except Exception as e:
        print("❌ Error exportando chat:", e)
        return jsonify({"error": "Error interno"}), 500

    


@consultas_bp.route("/api/files/<int:file_id>/download", methods=["GET"])
def descargar_archivo(file_id):
    username = request.args.get("username")  # 👈 capturamos el username
    ip = request.remote_addr  # 👈 capturamos la IP

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # Obtener nombre y túnel del archivo
        cursor.execute("SELECT filename, tunnel_id FROM tunnel_files WHERE id = %s", (file_id,))
        archivo = cursor.fetchone()
        cursor.close()
        conn.close()

        if not archivo:
            return jsonify({"error": "Archivo no encontrado"}), 404

        carpeta_archivos = os.path.join(os.getcwd(), "uploads")
        ruta_absoluta = os.path.join(carpeta_archivos, archivo["filename"])

        if not os.path.exists(ruta_absoluta):
            return jsonify({"error": "Archivo físico no encontrado"}), 404

        # ✅ Registrar el log
        if username:
            registrar_log(
                usuario=username,
                accion=f"Descargó el archivo '{archivo['filename']}' del túnel ID {archivo['tunnel_id']}",
                modulo="Archivos",
                ip=ip
            )

        return send_file(
            ruta_absoluta,
            as_attachment=True,
            download_name=archivo["filename"]
        )

    except Exception as e:
        print("❌ Error al descargar archivo:", e)
        return jsonify({"error": "Error interno"}), 500


@consultas_bp.route("/api/logs", methods=["GET"])
def consultar_logs():
    username = request.args.get("username")  # quien consulta
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")

    if not username:
        return jsonify({"error": "Falta el nombre de usuario"}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # Verificar si tiene rol admin
        cursor.execute("SELECT rol FROM usuarios WHERE username = %s", (username,))
        result = cursor.fetchone()
        if not result or result["rol"] != "admin":
            return jsonify({"error": "No autorizado"}), 403

        # Consulta general de logs (sin filtro por usuario)
        query = "SELECT id, usuario, accion, modulo, fecha AS timestamp, ip FROM logs WHERE 1=1"
        params = []

        if desde:
            query += " AND fecha >= %s"
            params.append(int(datetime.strptime(desde, "%Y-%m-%d").timestamp() * 1000))
        if hasta:
            query += " AND fecha <= %s"
            params.append(int(datetime.strptime(hasta, "%Y-%m-%d").timestamp() * 1000) + 86399999)

        query += " ORDER BY fecha DESC LIMIT 500"

        cursor.execute(query, tuple(params))
        logs = cursor.fetchall()

        return jsonify(logs)
    except Exception as e:
        print("❌ Error consultando logs:", e)
        return jsonify({"error": "Error interno"}), 500
    finally:
        cursor.close()
        conn.close()




